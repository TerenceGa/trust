# calculation_logic.py
# Handles interaction with the Excel calculation toolbox using openpyxl and LibreOffice
# Updated read_results_from_xlsx for specific cell locations.
# ADDED MORE DEBUG PRINT STATEMENTS
# ADDED Formula reading and explicit cell checks
# UPDATED: Removed verbose st.info/st.write messages for cleaner UI. Kept st.error and critical st.warning.

import streamlit as st # For displaying messages/errors during calculation
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import os
import shutil
import subprocess
import platform
import time
import traceback # Import traceback for detailed error printing

# --- Configuration ---
CALCULATOR_SHEET_NAME = "TRST"
YEAR_TO_CELL_MAP = {
    10: 'G74', 15: 'G79', 20: 'G84', 25: 'G89', 30: 'G94',
    35: 'G99', 40: 'G104', 45: 'G109', 50: 'G114', 55: 'G119',
    60: 'G124', 70: 'G134', 80: 'G144', 90: 'G154'
}

# --- Helper Function to Find LibreOffice ---
# (Keep the previous version with debug prints)
def find_soffice_path():
    """Attempts to find the path to the LibreOffice executable."""
    print("--- DEBUG (calc_logic): Entering find_soffice_path ---") # DEBUG
    path = None
    try:
        if platform.system() == "Windows":
            print("--- DEBUG (calc_logic): Checking Windows paths... ---") # DEBUG
            prog_files = os.environ.get("ProgramFiles", "C:\\Program Files")
            prog_files_x86 = os.environ.get("ProgramFiles(x86)", "C:\\Program Files (x86)")
            paths_to_check = [
                os.path.join(prog_files, "LibreOffice", "program", "soffice.exe"),
                os.path.join(prog_files_x86, "LibreOffice", "program", "soffice.exe"),
            ]
            for p in paths_to_check:
                # print(f"--- DEBUG (calc_logic): Checking path: {p} ---") # DEBUG (Optional)
                if os.path.exists(p):
                    print(f"--- DEBUG (calc_logic): Found at: {p} ---") # DEBUG
                    path = p
                    break
            if not path:
                print("--- DEBUG (calc_logic): Checking 'where' command... ---") # DEBUG
                try:
                    # Use shell=True cautiously, ensure command is safe
                    result = subprocess.run(["where", "soffice.exe"], capture_output=True, text=True, check=True, shell=True)
                    paths = result.stdout.strip().splitlines()
                    if paths and os.path.exists(paths[0]):
                         print(f"--- DEBUG (calc_logic): Found via 'where': {paths[0]} ---") # DEBUG
                         path = paths[0]
                except (FileNotFoundError, subprocess.CalledProcessError) as where_err:
                    print(f"--- DEBUG (calc_logic): 'where' command failed or not found: {where_err} ---") # DEBUG
                    pass # Not found via 'where'
        elif platform.system() == "Darwin": # macOS
             print("--- DEBUG (calc_logic): Checking macOS paths... ---")
             p = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
             if os.path.exists(p):
                  print(f"--- DEBUG (calc_logic): Found at: {p} ---")
                  path = p
             if not path: # Try 'which' command as a fallback
                try:
                    result = subprocess.run(["which", "soffice"], capture_output=True, text=True, check=True)
                    p_which = result.stdout.strip()
                    if p_which and os.path.exists(p_which):
                        print(f"--- DEBUG (calc_logic): Found via 'which': {p_which} ---")
                        path = p_which
                except (FileNotFoundError, subprocess.CalledProcessError) as e:
                    print(f"--- DEBUG (calc_logic): 'which soffice' failed: {e} ---")
        else: # Linux/Other
             print("--- DEBUG (calc_logic): Checking Linux/Other paths using 'which'... ---")
             try:
                result = subprocess.run(["which", "soffice"], capture_output=True, text=True, check=True)
                p_which = result.stdout.strip()
                if p_which and os.path.exists(p_which):
                    print(f"--- DEBUG (calc_logic): Found via 'which': {p_which} ---")
                    path = p_which
                else: # Check common install locations if 'which' fails or path is invalid
                    common_paths = ["/usr/bin/soffice", "/usr/local/bin/soffice", "/opt/libreoffice/program/soffice", "/snap/bin/libreoffice.soffice"]
                    for p_common in common_paths:
                        if os.path.exists(p_common):
                            print(f"--- DEBUG (calc_logic): Found in common path: {p_common} ---")
                            path = p_common
                            break
             except (FileNotFoundError, subprocess.CalledProcessError) as e:
                print(f"--- DEBUG (calc_logic): 'which soffice' failed: {e}. Will check common paths. ---")
                common_paths = ["/usr/bin/soffice", "/usr/local/bin/soffice", "/opt/libreoffice/program/soffice", "/snap/bin/libreoffice.soffice"]
                for p_common in common_paths:
                    if os.path.exists(p_common):
                        print(f"--- DEBUG (calc_logic): Found in common path: {p_common} ---")
                        path = p_common
                        break
    except Exception as e:
        print(f"--- ERROR (calc_logic): Exception in find_soffice_path: {e} ---") # DEBUG
        print(traceback.format_exc())
        path = None

    print(f"--- DEBUG (calc_logic): Exiting find_soffice_path. Found path: {path} ---") # DEBUG
    return path


# --- UPDATED Function to Read Results ---
def read_results_from_xlsx(calculated_xlsx_path, report_years):
    """
    Reads the calculated results from the specified XLSX file for the required years.
    Uses the YEAR_TO_CELL_MAP to read from specific cells.
    Also reads the formula from G74 and checks C37/AH74 directly.
    """
    results = []
    print(f"--- DEBUG (calc_logic): Entering read_results_from_xlsx for: {calculated_xlsx_path} ---") # DEBUG
    # st.info(f"Attempting to read results from specific cells in: {calculated_xlsx_path}") # Removed UI print
    if not os.path.exists(calculated_xlsx_path):
        print(f"--- ERROR (calc_logic): Calculated file not found: {calculated_xlsx_path} ---") # DEBUG
        st.error(f"計算後的檔案未找到: {calculated_xlsx_path}") # Keep UI error
        return None

    workbook = None # Initialize workbook to None
    try:
        # --- Load workbook WITH formulas first to check G74 ---
        print(f"--- DEBUG (calc_logic): Loading workbook {calculated_xlsx_path} (data_only=False to read formula)... ---") # DEBUG
        workbook_formulas = openpyxl.load_workbook(calculated_xlsx_path, data_only=False)
        if CALCULATOR_SHEET_NAME not in workbook_formulas.sheetnames:
             print(f"--- ERROR (calc_logic): Sheet '{CALCULATOR_SHEET_NAME}' not found (formula check). ---") # DEBUG
             workbook_formulas.close()
             # Fallback or error? For now, let's try loading with data_only=True
        else:
            sheet_formulas = workbook_formulas[CALCULATOR_SHEET_NAME]
            try:
                g74_formula = sheet_formulas['G74'].value # If it's a formula, .value gives the string
                print(f"--- DEBUG (calc_logic): Formula read from G74: {g74_formula} (Type: {type(g74_formula)}) ---") # DEBUG
                if sheet_formulas['G74'].data_type == 'f':
                     print("--- DEBUG (calc_logic): G74 data type is 'f' (formula). ---")
                else:
                     print(f"--- DEBUG (calc_logic): G74 data type is '{sheet_formulas['G74'].data_type}'. ---")

                c37_val_f = sheet_formulas['C37'].value
                c37_dt_f = sheet_formulas['C37'].data_type
                print(f"--- DEBUG (calc_logic): Read from C37 (formulas): Value={c37_val_f}, Type={c37_dt_f} ---")
                ah74_val_f = sheet_formulas['AH74'].value
                ah74_dt_f = sheet_formulas['AH74'].data_type
                print(f"--- DEBUG (calc_logic): Read from AH74 (formulas): Value={ah74_val_f}, Type={ah74_dt_f} ---")

            except Exception as formula_read_err:
                print(f"--- WARNING (calc_logic): Could not read formula/cells G74/C37/AH74: {formula_read_err} ---") # DEBUG
            workbook_formulas.close() # Close the formula workbook

        # --- Now load workbook with data_only=True to get calculated values ---
        print(f"--- DEBUG (calc_logic): Loading workbook {calculated_xlsx_path} (data_only=True)... ---") # DEBUG
        workbook = openpyxl.load_workbook(calculated_xlsx_path, data_only=True)
        if CALCULATOR_SHEET_NAME not in workbook.sheetnames:
            print(f"--- ERROR (calc_logic): Result Sheet '{CALCULATOR_SHEET_NAME}' not found (data read). ---") # DEBUG
            st.error(f"結果工作表 '{CALCULATOR_SHEET_NAME}' 未在檔案 {os.path.basename(calculated_xlsx_path)} 中找到。") # Keep UI error
            workbook.close()
            return None
        sheet = workbook[CALCULATOR_SHEET_NAME]
        print(f"--- DEBUG (calc_logic): Successfully opened sheet '{CALCULATOR_SHEET_NAME}' for data reading. ---") # DEBUG
        # st.info(f"Successfully opened sheet '{CALCULATOR_SHEET_NAME}'.") # Removed UI print

        try:
             c37_val_d = sheet['C37'].value
             print(f"--- DEBUG (calc_logic): Read from C37 (data_only): Value={c37_val_d} ---")
             ah74_val_d = sheet['AH74'].value
             print(f"--- DEBUG (calc_logic): Read from AH74 (data_only): Value={ah74_val_d} ---")
        except Exception as data_read_err:
             print(f"--- WARNING (calc_logic): Could not read cells C37/AH74 (data_only): {data_read_err} ---") # DEBUG

        for year in report_years:
            if year in YEAR_TO_CELL_MAP:
                cell_ref = YEAR_TO_CELL_MAP[year]
                try:
                    cell_value = sheet[cell_ref].value
                    total_csv = float(cell_value) if cell_value is not None else 0.0
                    # print(f"--- DEBUG (calc_logic): Read year {year} from {cell_ref}: Value={cell_value}, Float={total_csv} ---") 
                except KeyError:
                     print(f"--- WARNING (calc_logic): Cell reference '{cell_ref}' for year {year} not found. ---") # DEBUG
                     st.warning(f"儲存格參考 '{cell_ref}' (年份 {year}) 未在工作表中找到。將使用 0.0。") # Keep UI warning
                     total_csv = 0.0
                except (ValueError, TypeError):
                    print(f"--- WARNING (calc_logic): Could not convert value '{cell_value}' from {cell_ref} (year {year}) to float. ---") # DEBUG
                    st.warning(f"無法將儲存格 {cell_ref} (年份 {year}) 的值 '{cell_value}' 轉換為數字。將使用 0.0。") # Keep UI warning
                    total_csv = 0.0
                results.append({'year': year, 'total_csv': round(total_csv, 2)})
            else:
                print(f"--- WARNING (calc_logic): Report year {year} not found in YEAR_TO_CELL_MAP. ---") # DEBUG
                st.warning(f"報告年份 {year} 未在預期儲存格映射中找到。將附加 0.0。") # Keep UI warning
                results.append({'year': year, 'total_csv': 0.0})

        workbook.close()
        print(f"--- DEBUG (calc_logic): Successfully read results for {len(results)} years. ---") # DEBUG
        # st.info(f"Successfully read results for {len(results)} years from specific cells.") # Removed UI print
        return results

    except Exception as e:
        print(f"--- ERROR (calc_logic): Error reading results: {e} ---") # DEBUG
        print(traceback.format_exc()) # Print detailed traceback to console
        st.error(f"讀取結果時發生錯誤 {calculated_xlsx_path}: {e}") # Keep UI error
        # st.error(traceback.format_exc()) # Avoid showing full traceback in UI
        if 'workbook_formulas' in locals() and workbook_formulas: workbook_formulas.close()
        if workbook: workbook.close()
        return None


# --- Main Function to Run a Calculation Scenario ---
def run_calculation_scenario(scenario_name, base_xlsx_path, temp_dir, scenario_params, cell_map, report_years):
    print(f"\n--- DEBUG (calc_logic): === ENTERING run_calculation_scenario for: {scenario_name} === ---")
    # st.info(f"--- Processing Scenario: {scenario_name} ---") # Removed UI print
    safe_scenario_name = "".join(c for c in scenario_name if c.isalnum() or c in (' ', '_')).rstrip().lower().replace(' ', '_')
    input_temp_xlsx_path = os.path.join(temp_dir, f"input_{safe_scenario_name}.xlsx")
    intermediate_ods_path = os.path.join(temp_dir, f"intermediate_{safe_scenario_name}.ods")
    calculated_temp_xlsx_path = os.path.join(temp_dir, f"calculated_{safe_scenario_name}.xlsx")
    libre_output_dir = temp_dir

    workbook = None 
    try:
        print(f"--- DEBUG (calc_logic): Step 1: Copying and Writing Inputs for {scenario_name} ---") # DEBUG
        if not os.path.exists(base_xlsx_path):
             print(f"--- ERROR (calc_logic): Base calculator file not found: {base_xlsx_path} ---") # DEBUG
             st.error(f"基礎計算機檔案未找到: {base_xlsx_path}") # Keep UI error
             return None
        print(f"--- DEBUG (calc_logic): Copying template '{os.path.basename(base_xlsx_path)}' to '{os.path.basename(input_temp_xlsx_path)}' ---") # DEBUG
        # st.info(f"Copying template '{os.path.basename(base_xlsx_path)}' to '{os.path.basename(input_temp_xlsx_path)}'") # Removed UI print
        shutil.copyfile(base_xlsx_path, input_temp_xlsx_path)

        print("--- DEBUG (calc_logic): Writing inputs to spreadsheet... ---") # DEBUG
        # st.info("Writing inputs...") # Removed UI print
        workbook = openpyxl.load_workbook(input_temp_xlsx_path)
        if CALCULATOR_SHEET_NAME not in workbook.sheetnames:
            print(f"--- ERROR (calc_logic): Sheet '{CALCULATOR_SHEET_NAME}' not found. ---") # DEBUG
            st.error(f"工作表 '{CALCULATOR_SHEET_NAME}' 未在 {input_temp_xlsx_path} 中找到。") # Keep UI error
            workbook.close()
            return None
        sheet = workbook[CALCULATOR_SHEET_NAME]

        for param_name, cell_ref in cell_map.items():
            if param_name in scenario_params:
                try:
                    value_to_write = scenario_params[param_name]
                    if isinstance(value_to_write, (int, float)):
                        sheet[cell_ref].value = value_to_write
                    else:
                        sheet[cell_ref] = str(value_to_write)
                except Exception as write_err:
                     print(f"--- WARNING (calc_logic): Could not write '{param_name}' to cell {cell_ref}: {write_err} ---") # DEBUG
                     st.warning(f"無法將 '{param_name}' 寫入儲存格 {cell_ref}: {write_err}") # Keep UI warning

        workbook.save(input_temp_xlsx_path)
        workbook.close()
        workbook = None 
        print("--- DEBUG (calc_logic): Inputs written successfully. ---") # DEBUG
        # st.info("Inputs written successfully.") # Removed UI print

    except Exception as e:
        print(f"--- ERROR (calc_logic): Error preparing input file: {e} ---") # DEBUG
        print(traceback.format_exc()) 
        st.error(f"準備輸入檔案時發生錯誤 ({scenario_name}): {e}") # Keep UI error
        # st.error(traceback.format_exc()) # Avoid full traceback in UI
        if workbook: workbook.close()
        return None

    print(f"--- DEBUG (calc_logic): Step 2: Triggering Calculation for {scenario_name} ---") # DEBUG
    soffice_path = find_soffice_path()
    if not soffice_path:
        print("--- ERROR (calc_logic): LibreOffice 'soffice' command not found. Cannot proceed. ---") # DEBUG
        st.error("找不到 LibreOffice 'soffice' 命令。無法觸發計算。請確保已安裝 LibreOffice 並將其添加到系統 PATH 或標準位置。") # Keep UI error
        return None

    if os.path.exists(intermediate_ods_path):
        try: os.remove(intermediate_ods_path)
        except OSError as e: print(f"Warning: Could not remove old ODS file: {e}")
    if os.path.exists(calculated_temp_xlsx_path):
        try: os.remove(calculated_temp_xlsx_path)
        except OSError as e: print(f"Warning: Could not remove old XLSX file: {e}")

    ods_conversion_successful = False
    soffice_command_ods = [ soffice_path, "--headless", "--invisible", "--nologo", "--convert-to", "ods", "--outdir", libre_output_dir, input_temp_xlsx_path ]
    try:
        print(f"--- DEBUG (calc_logic): Running LibreOffice command (XLSX -> ODS): {' '.join(soffice_command_ods)} ---") # DEBUG
        # st.info(f"Running LibreOffice (Step 1/2: XLSX -> ODS): {' '.join(soffice_command_ods)}") # Removed UI print
        timeout_seconds = 90
        result_ods = subprocess.run(soffice_command_ods, capture_output=True, text=True, encoding='utf-8', errors='replace', timeout=timeout_seconds, check=False)
        print(f"--- DEBUG (calc_logic): LibreOffice (XLSX->ODS) finished. Return Code: {result_ods.returncode} ---")
        print(f"--- DEBUG (calc_logic): LibreOffice (XLSX->ODS) stdout:\n{result_ods.stdout or 'None'}\n---")
        print(f"--- DEBUG (calc_logic): LibreOffice (XLSX->ODS) stderr:\n{result_ods.stderr or 'None'}\n---")

        expected_ods_filename = os.path.splitext(os.path.basename(input_temp_xlsx_path))[0] + ".ods"
        expected_ods_path = os.path.join(libre_output_dir, expected_ods_filename)
        print(f"--- DEBUG (calc_logic): Expecting intermediate ODS file at: {expected_ods_path} ---")
        time.sleep(1) 

        if result_ods.returncode == 0 and os.path.exists(expected_ods_path):
            print(f"--- DEBUG (calc_logic): Intermediate ODS conversion successful. File at: {expected_ods_path} ---")
            os.rename(expected_ods_path, intermediate_ods_path)
            print(f"--- DEBUG (calc_logic): Renamed intermediate file to: {intermediate_ods_path} ---")
            ods_conversion_successful = True
        else:
            st.error(f"LibreOffice 在 XLSX -> ODS 轉換期間失敗 (RC={result_ods.returncode})。無法繼續。") # Keep UI error
            st.error(f"Stderr: {result_ods.stderr or 'None'}") # Keep UI error
            return None

    except Exception as e:
        print(f"--- ERROR (calc_logic): Error during XLSX -> ODS conversion: {e} ---") # DEBUG
        print(traceback.format_exc())
        st.error(f"LibreOffice XLSX->ODS 執行期間發生錯誤 ({scenario_name}): {e}") # Keep UI error
        # st.error(traceback.format_exc()) # Avoid full traceback in UI
        return None

    if not ods_conversion_successful: return None

    xlsx_conversion_successful = False
    soffice_command_xlsx = [ soffice_path, "--headless", "--invisible", "--nologo", "--convert-to", "xlsx", "--outdir", libre_output_dir, intermediate_ods_path ]
    try:
        print(f"--- DEBUG (calc_logic): Running LibreOffice command (ODS -> XLSX): {' '.join(soffice_command_xlsx)} ---") # DEBUG
        # st.info(f"Running LibreOffice (Step 2/2: ODS -> XLSX): {' '.join(soffice_command_xlsx)}") # Removed UI print
        timeout_seconds = 60
        result_xlsx = subprocess.run(soffice_command_xlsx, capture_output=True, text=True, encoding='utf-8', errors='replace', timeout=timeout_seconds, check=False)
        print(f"--- DEBUG (calc_logic): LibreOffice (ODS->XLSX) finished. Return Code: {result_xlsx.returncode} ---")
        print(f"--- DEBUG (calc_logic): LibreOffice (ODS->XLSX) stdout:\n{result_xlsx.stdout or 'None'}\n---")
        print(f"--- DEBUG (calc_logic): LibreOffice (ODS->XLSX) stderr:\n{result_xlsx.stderr or 'None'}\n---")

        expected_xlsx_filename = os.path.splitext(os.path.basename(intermediate_ods_path))[0] + ".xlsx"
        expected_xlsx_path = os.path.join(libre_output_dir, expected_xlsx_filename)
        print(f"--- DEBUG (calc_logic): Expecting final XLSX file at: {expected_xlsx_path} ---")
        time.sleep(1) 

        if result_xlsx.returncode == 0 and os.path.exists(expected_xlsx_path):
            print(f"--- DEBUG (calc_logic): Final XLSX conversion successful. File at: {expected_xlsx_path} ---")
            os.rename(expected_xlsx_path, calculated_temp_xlsx_path)
            print(f"--- DEBUG (calc_logic): Renamed final calculated file to: {calculated_temp_xlsx_path} ---")
            xlsx_conversion_successful = True
        else:
            st.error(f"LibreOffice 在 ODS -> XLSX 轉換期間失敗 (RC={result_xlsx.returncode})。無法繼續。") # Keep UI error
            st.error(f"Stderr: {result_xlsx.stderr or 'None'}") # Keep UI error
            return None

    except Exception as e:
        print(f"--- ERROR (calc_logic): Error during ODS -> XLSX conversion: {e} ---") # DEBUG
        print(traceback.format_exc())
        st.error(f"LibreOffice ODS->XLSX 執行期間發生錯誤 ({scenario_name}): {e}") # Keep UI error
        # st.error(traceback.format_exc()) # Avoid full traceback in UI
        return None

    try:
        if os.path.exists(intermediate_ods_path):
            os.remove(intermediate_ods_path)
            print(f"--- DEBUG (calc_logic): Removed intermediate ODS file: {intermediate_ods_path} ---")
    except Exception as e:
        print(f"--- WARNING (calc_logic): Could not remove intermediate ODS file: {e} ---")

    if not xlsx_conversion_successful:
        print(f"--- ERROR (calc_logic): Final XLSX conversion failed, cannot read results. ---") # DEBUG
        return None

    print(f"--- DEBUG (calc_logic): Step 3: Reading Results for {scenario_name} ---") # DEBUG
    print(f"--- DEBUG (calc_logic): Reading results from final calculated file: {calculated_temp_xlsx_path} ---")
    # st.info(f"Reading results from calculated file: {calculated_temp_xlsx_path}") # Removed UI print
    final_results_list = read_results_from_xlsx(calculated_temp_xlsx_path, report_years)

    print(f"--- DEBUG (calc_logic): === EXITING run_calculation_scenario for: {scenario_name}. Returning results: {'Success' if final_results_list else 'Failure/None'} === ---") # DEBUG
    return final_results_list
